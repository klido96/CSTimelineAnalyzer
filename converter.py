import csv
import sys

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

wb = openpyxl.Workbook()
ws = wb.active
# Timestamp, TreeID, PatternID, Parent Process, Process ID, Event, etc.
cell_width = [20, 12, 12, 15, 15, 35, 40, 30, 15, 40, 20, 15, 17, 35]


# Convert the csv to xlsx to work with openpyxl
def csv_to_xlsx(fn):
    with open(fn, encoding='utf8') as file:
        reader = csv.reader(file)
        for count, row in enumerate(reader):
            ws.append(row)
            ws.row_dimensions[count + 1].height = 15


# Apply filter, freeze the top row and format cell width
def filter_and_freeze_format():
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = 'A2'
    for count, col in enumerate(cell_width, start=1):
        ws.column_dimensions[get_column_letter(count)].width = cell_width[count - 1]
    # wrap text to avoid cell spilling. Will change to a toggle to reduce time
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrapText=True)

def main(filepath):
    try:
        csv_to_xlsx(filepath)
        filter_and_freeze_format()
        folder = '/'.join(filepath.split('/')[:-1]) + '/'
        output_file = filepath.split('/')[-1].split('.')[0]
        output_path = f'{folder}new_{output_file}.xlsx'
        wb.save(output_path)
        return output_path
    except:
        return 1


